"""Parse Excel workbooks into :class:`StandardDocument`."""

from __future__ import annotations

from pathlib import Path

from ee_wiki.common.config import ExcelConfig
from ee_wiki.common.errors import EEWikiError
from ee_wiki.common.logging import get_logger
from ee_wiki.common.types import DataLayoutConfig, Metadata, StandardDocument
from ee_wiki.ingestion.path_metadata import parse_path_metadata

logger = get_logger(__name__)

EXCEL_SUFFIXES = {".xlsx"}


class ExcelParserError(EEWikiError):
    """Failed to parse an Excel source file."""


def _cell_str(value: object) -> str:
    """Convert a workbook cell value to a plain string."""
    if value is None:
        return ""
    return str(value).strip()


def _escape_md_cell(text: str) -> str:
    """Escape Markdown table cell content."""
    return text.replace("|", "\\|").replace("\n", " ")


def _normalize_rows(rows: list[list[str]]) -> list[list[str]]:
    """Drop trailing empty rows and pad ragged rows to a common width."""
    trimmed = [row for row in rows if any(cell for cell in row)]
    if not trimmed:
        return []
    width = max(len(row) for row in trimmed)
    return [row + [""] * (width - len(row)) for row in trimmed]


def _sheet_rows(
    worksheet,
    *,
    max_rows: int | None,
) -> list[list[str]]:
    """Read worksheet rows as string grids."""
    rows: list[list[str]] = []
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if max_rows is not None and row_index > max_rows:
            logger.info(
                "Truncating sheet %s at %d row(s)",
                worksheet.title,
                max_rows,
            )
            break
        rows.append([_cell_str(value) for value in row])
    return _normalize_rows(rows)


def _rows_to_markdown_table(rows: list[list[str]]) -> str:
    """Render rows as a Markdown pipe table using the first row as header."""
    if not rows:
        return "_Empty sheet_\n"
    header = rows[0]
    body = rows[1:] if len(rows) > 1 else []
    header_line = "| " + " | ".join(_escape_md_cell(cell) for cell in header) + " |"
    separator = "| " + " | ".join("---" for _ in header) + " |"
    lines = [header_line, separator]
    for row in body:
        padded = row + [""] * (len(header) - len(row))
        lines.append(
            "| " + " | ".join(_escape_md_cell(cell) for cell in padded[: len(header)]) + " |"
        )
    return "\n".join(lines) + "\n"


def _rows_to_plain_text(rows: list[list[str]]) -> str:
    """Render rows as tab-separated plain text."""
    if not rows:
        return "_Empty sheet_\n"
    return "\n".join("\t".join(row) for row in rows) + "\n"


def _render_sheet(rows: list[list[str]], *, output_format: str) -> str:
    """Format one worksheet's rows according to ``output_format``."""
    if output_format == "plain_text":
        return _rows_to_plain_text(rows)
    if output_format != "markdown_table":
        raise ExcelParserError(f"Unsupported excel output_format: {output_format}")
    return _rows_to_markdown_table(rows)


def parse_excel(
    raw_path: Path,
    layout: DataLayoutConfig,
    excel_config: ExcelConfig,
    *,
    repo_root: Path | None = None,
    metadata: Metadata | None = None,
) -> StandardDocument:
    """Read an Excel workbook and build a :class:`StandardDocument`.

    Args:
        raw_path: Path to a ``.xlsx`` file. Normally under ``layout.raw_dir``; when
            ``metadata`` is provided (e.g. Numbers export temp file), path-derived
            metadata is taken from ``metadata`` instead.
        layout: Data layout configuration for path-derived metadata.
        excel_config: Excel ingest settings.
        repo_root: Optional repository root for ``source_file`` labels.
        metadata: Optional pre-parsed metadata when ``raw_path`` is not under
            ``layout.raw_dir`` (e.g. temporary export from ``.numbers``).

    Returns:
        Parsed document with one section per worksheet.

    Raises:
        ExcelParserError: If the workbook cannot be read or parsed.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExcelParserError(
            "openpyxl is required for Excel ingest; install with pip install 'ee-wiki[ml]'"
        ) from exc

    if metadata is None:
        metadata = parse_path_metadata(raw_path, layout, repo_root=repo_root)
    try:
        workbook = load_workbook(raw_path, read_only=True, data_only=True)
    except OSError as exc:
        raise ExcelParserError(f"Cannot read Excel file: {raw_path}") from exc
    except Exception as exc:
        raise ExcelParserError(f"Failed to open Excel workbook: {raw_path}") from exc

    sections: list[str] = [f"# {metadata.title}\n"]
    sheet_count = len(workbook.sheetnames)
    try:
        for worksheet in workbook.worksheets:
            rows = _sheet_rows(worksheet, max_rows=excel_config.max_rows_per_sheet)
            if not rows and not excel_config.include_empty_sheets:
                logger.info("Skipping empty Excel sheet: %s", worksheet.title)
                continue
            sections.append(f"## Sheet: {worksheet.title}\n")
            sections.append(
                _render_sheet(rows, output_format=excel_config.output_format)
            )
    finally:
        workbook.close()

    if len(sections) == 1:
        raise ExcelParserError(f"Excel workbook contains no ingestible sheets: {raw_path}")

    content = "\n".join(sections).rstrip() + "\n"
    document = StandardDocument(
        content=content,
        metadata=metadata,
        source_ref=str(raw_path.resolve()),
    )
    logger.info(
        "Parsed Excel %s (%d chars, sheets=%d, format=%s)",
        metadata.source_file,
        len(document.content),
        sheet_count,
        excel_config.output_format,
    )
    return document
