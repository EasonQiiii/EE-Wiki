"""Tests for Excel workbook ingest."""

from __future__ import annotations

from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook

from ee_wiki.common.config import AppConfig, ExcelConfig
from ee_wiki.ingestion.parsers.excel import ExcelParserError, parse_excel
from ee_wiki.ingestion.pipeline import ingest_file


@pytest.fixture
def excel_config(app_config, tmp_path: Path) -> AppConfig:
    raw_dir = tmp_path / "raw"
    processed_dir = tmp_path / "processed"
    raw_dir.mkdir()
    processed_dir.mkdir()
    layout = replace(
        app_config.data_layout,
        raw_dir=raw_dir,
        processed_dir=processed_dir,
    )
    return replace(
        app_config,
        raw_dir=raw_dir,
        processed_dir=processed_dir,
        data_layout=layout,
        repo_root=app_config.repo_root,
    )


def _write_sample_workbook(path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Part", "Qty"])
    summary.append(["R100", "10"])
    summary.append(["C200", "5"])

    parts = workbook.create_sheet("Parts")
    parts.append(["Ref", "Value", "描述"])
    parts.append(["U1", "TPS62840", "PMIC"])
    parts.append(["R1", "10k", "上拉"])

    workbook.create_sheet("Empty")
    workbook.save(path)
    workbook.close()


def test_parse_excel_markdown_table(excel_config: AppConfig) -> None:
    raw_path = excel_config.raw_dir / "iphone/logan/p1/note/bom.xlsx"
    raw_path.parent.mkdir(parents=True)
    _write_sample_workbook(raw_path)

    document = parse_excel(
        raw_path,
        excel_config.data_layout,
        ExcelConfig(output_format="markdown_table"),
        repo_root=excel_config.repo_root,
    )

    assert document.metadata.project == "logan"
    assert document.metadata.build == "p1"
    assert document.metadata.document_type == "engineering_note"
    assert "## Sheet: Summary" in document.content
    assert "## Sheet: Parts" in document.content
    assert "| Part | Qty |" in document.content
    assert "TPS62840" in document.content
    assert "PMIC" in document.content
    assert "## Sheet: Empty" not in document.content


def test_parse_excel_plain_text(excel_config: AppConfig) -> None:
    raw_path = excel_config.raw_dir / "iphone/logan/p1/note/bom.xlsx"
    raw_path.parent.mkdir(parents=True)
    _write_sample_workbook(raw_path)

    document = parse_excel(
        raw_path,
        excel_config.data_layout,
        ExcelConfig(output_format="plain_text"),
        repo_root=excel_config.repo_root,
    )

    assert "Part\tQty" in document.content
    assert "Ref\tValue\t描述" in document.content
    assert "| Part |" not in document.content


def test_ingest_file_writes_excel_mirror(excel_config: AppConfig) -> None:
    raw_path = excel_config.raw_dir / "iphone/logan/p1/note/bom.xlsx"
    raw_path.parent.mkdir(parents=True)
    _write_sample_workbook(raw_path)

    result = ingest_file(raw_path, excel_config)
    assert result.processed.content_path.suffix == ".md"
    assert result.processed.content_path.is_file()
    assert result.processed.metadata_path.is_file()
    assert "TPS62840" in result.processed.content_path.read_text(encoding="utf-8")


def test_parse_excel_rejects_workbook_with_only_empty_sheets(excel_config: AppConfig) -> None:
    raw_path = excel_config.raw_dir / "iphone/logan/p1/note/empty.xlsx"
    raw_path.parent.mkdir(parents=True)
    workbook = Workbook()
    workbook.save(raw_path)
    workbook.close()

    with pytest.raises(ExcelParserError, match="no ingestible sheets"):
        parse_excel(
            raw_path,
            excel_config.data_layout,
            ExcelConfig(),
            repo_root=excel_config.repo_root,
        )
